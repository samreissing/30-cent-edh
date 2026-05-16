"""Read the source Excel workbook and emit normalized JSON for the rest of the pipeline.

This is the "snapshot" path. The nightly GitHub Action will instead pull the live
Google Sheet, but the Excel parser is what we use locally and as a fallback.
"""

from __future__ import annotations

import json
from datetime import datetime, date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "20 word 30 Cent EDH.xlsx"


def _val(cell):
    if isinstance(cell, (datetime, date)):
        return cell.date().isoformat() if isinstance(cell, datetime) else cell.isoformat()
    return cell


def _get(row: tuple, idx: int):
    return row[idx] if idx < len(row) else None


def parse() -> dict:
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)

    master_names: list[str] = []
    ws = wb["Master Sheet"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        name = row[0]
        if name and isinstance(name, str) and name.strip():
            master_names.append(name.strip())

    illegal: list[dict] = []
    ws = wb["Illegal Cards"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row or not _get(row, 1):
            continue
        illegal.append({
            "date_checked": _val(_get(row, 0)),
            "name": (_get(row, 1) or "").strip(),
            "checked_by": (_get(row, 2) or "").strip() if _get(row, 2) else None,
            "scryfall_link": _get(row, 3),
            "status": _get(row, 4),
            "note": _get(row, 5),
        })

    watchlist: list[dict] = []
    ws = wb["Watchlist"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row or not _get(row, 1):
            continue
        watchlist.append({
            "date_checked": _val(_get(row, 0)),
            "name": (_get(row, 1) or "").strip(),
            "checked_by": (_get(row, 2) or "").strip() if _get(row, 2) else None,
            "scryfall_link": _get(row, 3),
            "status": _get(row, 4),
        })

    # The Master Sheet has dupes (a card can appear in multiple tabs). Dedupe.
    seen = set()
    deduped = []
    for n in master_names:
        key = n.lower()
        if key in seen:
            continue
        seen.add(key)
        deduped.append(n)

    return {
        "source": "excel_snapshot",
        "workbook": WORKBOOK.name,
        "cardpool": deduped,
        "banlist": illegal,
        "watchlist": watchlist,
    }


def main() -> None:
    out_dir = ROOT / "data"
    out_dir.mkdir(exist_ok=True)
    data = parse()
    cardpool_path = out_dir / "cardpool.json"
    banlist_path = out_dir / "banlist.json"
    watchlist_path = out_dir / "watchlist.json"
    cardpool_path.write_text(json.dumps(data["cardpool"], indent=2))
    banlist_path.write_text(json.dumps(data["banlist"], indent=2))
    watchlist_path.write_text(json.dumps(data["watchlist"], indent=2))
    print(f"cardpool : {len(data['cardpool']):>6} cards -> {cardpool_path}")
    print(f"banlist  : {len(data['banlist']):>6} cards -> {banlist_path}")
    print(f"watchlist: {len(data['watchlist']):>6} cards -> {watchlist_path}")


if __name__ == "__main__":
    main()
